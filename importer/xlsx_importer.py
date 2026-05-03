# -*- coding: utf-8 -*-
"""
importer/xlsx_importer.py — Excel 文件题库解析器

支持的 Excel 列格式（第一行为表头，从第二行开始是数据）：
  A: 题型   (single/multi/truefalse/scenario 或 单选/多选/判断/场景)
  B: 题干
  C: 选项A
  D: 选项B
  E: 选项C  (可选)
  F: 选项D  (可选)
  G: 选项E  (可选)
  H: 正确答案
  I: 解析   (可选)
  J: 分类   (可选)
  K: 难度   (1/2/3，可选)

多个 Sheet 代表不同分类，Sheet 名称作为默认分类（若 J 列为空则使用 Sheet 名）。
"""

import logging
from typing import Any, Dict, List, Optional

from importer.base_importer import BaseImporter

logger = logging.getLogger(__name__)

# 中文题型名称 → 内部标识符
TYPE_MAP = {
    'single':    'single',
    'multi':     'multi',
    'truefalse': 'truefalse',
    'scenario':  'scenario',
    '单选':      'single',
    '单选题':    'single',
    '多选':      'multi',
    '多选题':    'multi',
    '判断':      'truefalse',
    '判断题':    'truefalse',
    '场景':      'scenario',
    '场景题':    'scenario',
    '情景':      'scenario',
    '情景题':    'scenario',
}


class XlsxImporter(BaseImporter):
    """Excel (.xlsx) 题库导入器"""

    def parse(self) -> List[Dict[str, Any]]:
        """解析 Excel 文件，返回题目字典列表"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError('缺少依赖 openpyxl，请运行: pip install openpyxl')

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        questions: List[Dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_questions = self._parse_sheet(ws, sheet_name)
            questions.extend(sheet_questions)
            logger.info('Sheet "%s" 解析完成，发现 %d 题', sheet_name, len(sheet_questions))

        wb.close()
        logger.info('Excel 解析完成，共发现 %d 题', len(questions))
        return questions

    def _parse_sheet(self, ws, sheet_name: str) -> List[Dict[str, Any]]:
        """解析单个 Sheet"""
        questions: List[Dict[str, Any]] = []
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return questions

        # 跳过表头行（第一行）
        data_rows = rows[1:]

        for row_idx, row in enumerate(data_rows, start=2):
            # 补齐列数，避免短行越界
            row = list(row) + [None] * 11
            q = self._parse_row(row, sheet_name, row_idx)
            if q:
                questions.append(q)

        return questions

    def _parse_row(self, row: list, sheet_name: str, row_idx: int) -> Optional[Dict[str, Any]]:
        """解析单行数据，返回题目字典或 None"""

        def cell(idx: int) -> str:
            """安全读取单元格，转为字符串并去除首尾空白"""
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else ''

        q_type_raw = cell(0)
        content    = cell(1)
        option_a   = cell(2)
        option_b   = cell(3)
        option_c   = cell(4)
        option_d   = cell(5)
        option_e   = cell(6)
        correct    = cell(7)
        explanation= cell(8)
        category   = cell(9) or sheet_name   # J 列为空时用 Sheet 名
        difficulty_raw = cell(10)

        # 跳过空行
        if not content and not correct:
            return None

        # 解析题型
        q_type = TYPE_MAP.get(q_type_raw.lower(), TYPE_MAP.get(q_type_raw, ''))
        if not q_type:
            # 无法识别题型时，根据答案内容猜测
            q_type = self._guess_type(correct, option_a, option_b)
            logger.debug('第 %d 行题型 "%s" 无法识别，猜测为 %s', row_idx, q_type_raw, q_type)

        # 解析难度
        try:
            difficulty = int(difficulty_raw) if difficulty_raw in ('1', '2', '3') else 2
        except (ValueError, TypeError):
            difficulty = 2

        return {
            'q_type':      q_type,
            'content':     content,
            'option_a':    option_a,
            'option_b':    option_b,
            'option_c':    option_c,
            'option_d':    option_d,
            'option_e':    option_e,
            'correct_ans': correct,
            'explanation': explanation,
            'category':    category,
            'difficulty':  difficulty,
        }

    @staticmethod
    def _guess_type(correct: str, option_a: str, option_b: str) -> str:
        """根据答案和选项内容猜测题型"""
        c = correct.strip().lower()
        if c in ('true', 'false', '正确', '错误', '对', '错', '√', '×'):
            return 'truefalse'
        if len(c) > 1 and all(ch.isalpha() for ch in c):
            return 'multi'
        if not option_a and not option_b:
            return 'truefalse'
        return 'single'
